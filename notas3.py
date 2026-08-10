# -*- coding: utf-8 -*-
"""Extrai notas tratando os blocos em cache desatualizado da planilha.
Regras: (1) a aba ESTUDANTES é a lista oficial de matrícula;
(2) quando uma turma aparece em mais de um bloco no mesmo período, vence o bloco
    mais consistente com a lista oficial (menos nomes estranhos à turma);
(3) só entram estudantes que constam na lista oficial da turma."""
import openpyxl, json, re, collections
SRC='/sessions/amazing-affectionate-davinci/mnt/uploads/PROFESSORA MARIA CASIMIRO SOARES - REGISTRO DE NOTAS DOS ESTUDANTES - 2026 (1).xlsx'
POS=list(range(38,227,7)); PER=['1º Período','2º Período','3º Período','4º Período']
wb=openpyxl.load_workbook(SRC,data_only=True,read_only=True)
def num(v): return round(float(v),2) if isinstance(v,(int,float)) else None
def txt(v): return re.sub(r'\s+',' ',str(v)).strip() if v is not None else ''
def lab(code):
    m=re.match(r'.*?-([A-Z]+)-(\d+)$',code); 
    if not m: return code
    return m.group(2) if m.group(1)=='EPT' else m.group(2)+' '+m.group(1)
def ser(code):
    m=re.search(r'-(\d)\d\d$',code); return {'1':'1º Ano','2':'2º Ano','3':'3º Ano'}.get(m.group(1),'—') if m else '—'

roster=collections.defaultdict(set)
for r in wb['ESTUDANTES'].iter_rows(min_row=2,values_only=True):
    if r and isinstance(r[2],str) and isinstance(r[3],str) and r[2].startswith('26PMI'):
        roster[txt(r[2])].add(txt(r[3]))

COMPS=[]; 
def cidx(c):
    if c not in COMPS: COMPS.append(c)
    return COMPS.index(c)
recs={}; qual=[]; usados={}
for pi,pname in enumerate(PER,1):
    ws=wb[pname]
    rows=[list(r)+[None]*(263-len(r)) for r in ws.iter_rows(min_row=1,max_row=2100,max_col=263,values_only=True)]
    hdr=[i for i,r in enumerate(rows) if txt(r[4])=='Turma']
    blocos=[]
    for bi,h in enumerate(hdr):
        e=hdr[bi+1] if bi+1<len(hdr) else len(rows)
        st=[r for r in rows[h:e] if isinstance(r[1],(int,float)) and isinstance(r[2],str)
            and isinstance(r[4],str) and txt(r[4]) not in ('NOME','Turma')]
        if not st: continue
        turma=txt(st[0][2]); nomes={txt(r[4]) for r in st}
        blocos.append({'l':h+1,'turma':turma,'st':st,'comp':[txt(rows[h][p]) for p in POS],
                       'extra':len(nomes-roster.get(turma,set())),'cob':len(nomes&roster.get(turma,set()))})
    best={}
    for b in blocos:
        k=b['turma']
        if k not in best or (b['extra'],-b['cob'])<(best[k]['extra'],-best[k]['cob']): best[k]=b
    for k,b in sorted(best.items()):
        desc=[x['l'] for x in blocos if x['turma']==k and x['l']!=b['l']]
        qual.append({'p':pi,'turma':lab(k),'linha':b['l'],'alunos':b['cob'],'extras':b['extra'],'descartados':desc})
        usados.setdefault(lab(k),set()).add(str(pi))
        for r in b['st']:
            nome=txt(r[4])
            if nome not in roster.get(k,set()): continue
            a=recs.setdefault((k,nome),{'t':lab(k),'s':ser(k),'n':nome,'st':{},'p':{}})
            if txt(r[36]): a['st'][str(pi)]=txt(r[36])
            d={}
            for ci,p in enumerate(POS):
                cn=re.sub(r'\s*-\s*$','',b['comp'][ci] or '').strip()
                if not cn or cn.upper() in ('FTP','APF','-'): continue
                vals=[num(r[p]),num(r[p+2]),num(r[p+4]),num(r[p+1]),num(r[p+3]),num(r[p+5]),num(r[p+6])]
                if all(v is None for v in vals): continue
                d[str(cidx(cn))]=vals
            if d: a['p'][str(pi)]=d
alunos=[v for v in recs.values() if v['p']]
alunos.sort(key=lambda a:(a['s'],a['t'],a['n']))
turmas=sorted({a['t'] for a in alunos})
falta=[{'turma':t,'p':p} for t in turmas for p in ['1','2'] if p not in usados.get(t,set())]
out={'comp':COMPS,'turmas':turmas,'series':sorted({a['s'] for a in alunos}),'per':PER,
     'perComNotas':sorted({k for a in alunos for k in a['p']}),'alunos':alunos,
     'qual':qual,'semDados':falta,
     'roster':{lab(k):len(v) for k,v in roster.items()}}
json.dump(out,open('notas.json','w',encoding='utf-8'),ensure_ascii=False,separators=(',',':'))
import os
print('alunos',len(alunos),'| turmas',len(turmas),'| json KB %.0f'%(os.path.getsize('notas.json')/1024))
print('\nblocos escolhidos:')
for q in qual: print('  P%d %-9s linha %-5d alunos=%-3d extras=%d descartados=%s'%(q['p'],q['turma'],q['linha'],q['alunos'],q['extras'],q['descartados'] or '-'))
print('\nsem lançamento:',falta)
print('\nconferência de contagem (alunos x lista oficial):')
for t in turmas:
    n=sum(1 for a in alunos if a['t']==t); print('  %-9s painel=%-3d oficial=%d'%(t,n,out['roster'][t]))
