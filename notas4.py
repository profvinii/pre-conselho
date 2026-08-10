# -*- coding: utf-8 -*-
"""Extrai notas + área do conhecimento (linha 7 da planilha, por posição de coluna).
Blocos em cache desatualizado: vence o mais consistente com a aba ESTUDANTES."""
import openpyxl, json, re, collections
SRC='/sessions/amazing-affectionate-davinci/mnt/uploads/PROFESSORA MARIA CASIMIRO SOARES - REGISTRO DE NOTAS DOS ESTUDANTES - 2026 (1).xlsx'
POS=list(range(38,227,7)); PER=['1º Período','2º Período','3º Período','4º Período']
AREAS=[(38,'Linguagens'),(73,'Humanas'),(101,'Natureza'),(122,'Matemática'),
       (129,'Formação Técnica e Profissional'),(171,'Letramento'),(185,'Aprofundamentos')]
ANOMES=[a[1] for a in AREAS]
def area_de(p):
    a=ANOMES[0]
    for col,nome in AREAS:
        if p>=col: a=nome
    return a
wb=openpyxl.load_workbook(SRC,data_only=True,read_only=True)
def num(v): return round(float(v),2) if isinstance(v,(int,float)) else None
def txt(v): return re.sub(r'\s+',' ',str(v)).strip() if v is not None else ''
def lab(c):
    m=re.match(r'.*?-([A-Z]+)-(\d+)$',c)
    return c if not m else (m.group(2) if m.group(1)=='EPT' else m.group(2)+' '+m.group(1))
def ser(c):
    m=re.search(r'-(\d)\d\d$',c); return {'1':'1º Ano','2':'2º Ano','3':'3º Ano'}.get(m.group(1),'—') if m else '—'
def nome_comp(bruto,p):
    n=re.sub(r'\s*-\s*$','',bruto or '').strip()
    a=area_de(p)
    if a=='Formação Técnica e Profissional' and (not n or n.upper()=='FTP'): return 'FTP · módulo %d'%((p-129)//7+1)
    if a=='Aprofundamentos' and (not n or re.fullmatch(r'APF ?\d*',n,re.I)): return 'Aprofundamento %d'%((p-185)//7+1)
    return n
roster=collections.defaultdict(set)
for r in wb['ESTUDANTES'].iter_rows(min_row=2,values_only=True):
    if r and isinstance(r[2],str) and isinstance(r[3],str) and r[2].startswith('26PMI'):
        roster[txt(r[2])].add(txt(r[3]))
COMPS=[]; CAREA=[]
def cidx(nome,p):
    if nome not in COMPS: COMPS.append(nome); CAREA.append(ANOMES.index(area_de(p)))
    return COMPS.index(nome)
recs={}; qual=[]; usados={}
for pi,pname in enumerate(PER,1):
    rows=[list(r)+[None]*(263-len(r)) for r in wb[pname].iter_rows(min_row=1,max_row=2100,max_col=263,values_only=True)]
    hdr=[i for i,r in enumerate(rows) if txt(r[4])=='Turma']; blocos=[]
    for bi,h in enumerate(hdr):
        e=hdr[bi+1] if bi+1<len(hdr) else len(rows)
        st=[r for r in rows[h:e] if isinstance(r[1],(int,float)) and isinstance(r[2],str)
            and isinstance(r[4],str) and txt(r[4]) not in ('NOME','Turma')]
        if not st: continue
        t=txt(st[0][2]); nm={txt(r[4]) for r in st}
        blocos.append({'l':h+1,'turma':t,'st':st,'hdr':rows[h],
                       'extra':len(nm-roster.get(t,set())),'cob':len(nm&roster.get(t,set()))})
    best={}
    for b in blocos:
        k=b['turma']
        if k not in best or (b['extra'],-b['cob'])<(best[k]['extra'],-best[k]['cob']): best[k]=b
    for k,b in sorted(best.items()):
        qual.append({'p':pi,'turma':lab(k),'linha':b['l'],'alunos':b['cob'],'extras':b['extra'],
                     'descartados':[x['l'] for x in blocos if x['turma']==k and x['l']!=b['l']]})
        usados.setdefault(lab(k),set()).add(str(pi))
        for r in b['st']:
            n=txt(r[4])
            if n not in roster.get(k,set()): continue
            a=recs.setdefault((k,n),{'t':lab(k),'s':ser(k),'n':n,'st':{},'p':{}})
            if txt(r[36]): a['st'][str(pi)]=txt(r[36])
            d={}
            for p in POS:
                cn=nome_comp(txt(b['hdr'][p]),p)
                if not cn: continue
                vals=[num(r[p]),num(r[p+2]),num(r[p+4]),num(r[p+1]),num(r[p+3]),num(r[p+5]),num(r[p+6])]
                if all(v is None for v in vals): continue
                d[str(cidx(cn,p))]=vals
            if d and (str(pi) not in a['p'] or len(d)>len(a['p'][str(pi)])): a['p'][str(pi)]=d
alunos=[v for v in recs.values() if v['p']]
alunos.sort(key=lambda a:(a['s'],a['t'],a['n']))
turmas=sorted({a['t'] for a in alunos})
ordem={n:i for i,n in enumerate(ANOMES)}
out={'comp':COMPS,'compArea':CAREA,'areas':[n for n in ANOMES if n in {ANOMES[i] for i in CAREA}],
     'turmas':turmas,'series':sorted({a['s'] for a in alunos}),'per':PER,
     'perComNotas':sorted({k for a in alunos for k in a['p']}),'alunos':alunos,'qual':qual,
     'semDados':[{'turma':t,'p':p} for t in turmas for p in ['1','2'] if p not in usados.get(t,set())],
     'roster':{lab(k):len(v) for k,v in roster.items()}}
json.dump(out,open('notas.json','w',encoding='utf-8'),ensure_ascii=False,separators=(',',':'))
import os
print('alunos',len(alunos),'| json KB %.0f'%(os.path.getsize('notas.json')/1024))
print('\ncomponentes por área:')
for ai,an in enumerate(ANOMES):
    cs=[COMPS[i] for i in range(len(COMPS)) if CAREA[i]==ai]
    if cs: print('  %-32s %s'%(an,cs))
print('\náreas ativas:',out['areas'])
print('sem lançamento:',out['semDados'])
