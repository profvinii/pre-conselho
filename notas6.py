# -*- coding: utf-8 -*-
"""Notas + área do conhecimento (v6: FTP com o nome real da disciplina).
Tratamentos: (a) bloco em cache desatualizado -> vence o consistente com a aba ESTUDANTES;
(b) linha duplicada do mesmo estudante -> vence a mais completa (empate: a de menor nº);
(c) nota fora da escala -> limitada a 0..10, com registro da ocorrência."""
import openpyxl, json, re, collections
SRC='/sessions/amazing-affectionate-davinci/mnt/uploads/PROFESSORA MARIA CASIMIRO SOARES - REGISTRO DE NOTAS DOS ESTUDANTES - 2026 (1).xlsx'
POS=list(range(38,227,7)); PER=['1º Período','2º Período','3º Período','4º Período']
AREAS=[(38,'Linguagens'),(73,'Humanas'),(101,'Natureza'),(122,'Matemática'),
       (129,'Formação Técnica e Profissional'),(171,'Letramento'),(185,'Aprofundamentos')]
ANOMES=[a[1] for a in AREAS]
CAMPOS=['AV1','AV2','AV3','RP1','RP2','RP3','MÉDIA']
def area_de(p):
    a=ANOMES[0]
    for col,n in AREAS:
        if p>=col: a=n
    return a
wb=openpyxl.load_workbook(SRC,data_only=True,read_only=True)
def txt(v): return re.sub(r'\s+',' ',str(v)).strip() if v is not None else ''
def lab(c):
    m=re.match(r'.*?-([A-Z]+)-(\d+)$',c)
    return c if not m else (m.group(2) if m.group(1)=='EPT' else m.group(2)+' '+m.group(1))
def ser(c):
    m=re.search(r'-(\d)\d\d$',c); return {'1':'1º Ano','2':'2º Ano','3':'3º Ano'}.get(m.group(1),'—') if m else '—'
def nome_comp(b,p,abaixo=''):
    n=re.sub(r'\s*-\s*$','',b or '').strip(); a=area_de(p); ab=txt(abaixo)
    if a=='Formação Técnica e Profissional' and (not n or n.upper()=='FTP'):
        return ab if ab else 'FTP · módulo %d'%((p-129)//7+1)
    if a=='Aprofundamentos' and (not n or re.fullmatch(r'APF ?\d*',n,re.I)):
        return ab if ab else 'Aprofundamento %d'%((p-185)//7+1)
    return n
roster=collections.defaultdict(set)
for r in wb['ESTUDANTES'].iter_rows(min_row=2,values_only=True):
    if r and isinstance(r[2],str) and isinstance(r[3],str) and r[2].startswith('26PMI'):
        roster[txt(r[2])].add(txt(r[3]))
COMPS=[]; CAREA=[]; incons=[]; dups=[]
def cidx(n,p):
    if n not in COMPS: COMPS.append(n); CAREA.append(ANOMES.index(area_de(p)))
    return COMPS.index(n)
recs={}; qual=[]; usados={}
for pi,pname in enumerate(PER,1):
    rows=[list(r)+[None]*(263-len(r)) for r in wb[pname].iter_rows(min_row=1,max_row=2100,max_col=263,values_only=True)]
    hdr=[i for i,r in enumerate(rows) if txt(r[4])=='Turma']; blocos=[]
    for bi,h in enumerate(hdr):
        e=hdr[bi+1] if bi+1<len(hdr) else len(rows)
        st=[(i,r) for i,r in enumerate(rows[h:e],h) if isinstance(r[1],(int,float)) and isinstance(r[2],str)
            and isinstance(r[4],str) and txt(r[4]) not in ('NOME','Turma')]
        if not st: continue
        t=txt(st[0][1][2]); nm={txt(r[4]) for _,r in st}
        blocos.append({'l':h+1,'turma':t,'st':st,'hdr':rows[h],'sub':rows[h+1] if h+1<len(rows) else [None]*263,
                       'extra':len(nm-roster.get(t,set())),'cob':len(nm&roster.get(t,set()))})
    best={}
    for b in blocos:
        k=b['turma']
        if k not in best or (b['extra'],-b['cob'])<(best[k]['extra'],-best[k]['cob']): best[k]=b
    for k,b in sorted(best.items()):
        qual.append({'p':pi,'turma':lab(k),'linha':b['l'],'alunos':b['cob'],'extras':b['extra'],
                     'descartados':[x['l'] for x in blocos if x['turma']==k and x['l']!=b['l']]})
        usados.setdefault(lab(k),set()).add(str(pi))
        porNome=collections.defaultdict(list)
        for i,r in b['st']: porNome[txt(r[4])].append((i,r))
        for n,occ in porNome.items():
            if n not in roster.get(k,set()): continue
            if len(occ)>1:
                dups.append({'p':pi,'turma':lab(k),'nome':n,
                             'linhas':[i+1 for i,_ in occ],'numeros':[int(r[1]) for _,r in occ]})
                occ=sorted(occ,key=lambda x:(-sum(1 for p in POS if isinstance(x[1][p+6],(int,float))),x[1][1]))
            i,r=occ[0]
            a=recs.setdefault((k,n),{'t':lab(k),'s':ser(k),'n':n,'st':{},'p':{}})
            if txt(r[36]): a['st'][str(pi)]=txt(r[36])
            d={}
            for p in POS:
                cn=nome_comp(txt(b['hdr'][p]),p,b['sub'][p])
                if not cn: continue
                vals=[]
                for off in (0,2,4,1,3,5,6):
                    v=r[p+off]
                    if isinstance(v,(int,float)):
                        v=float(v)
                        if v>10 or v<0:
                            incons.append({'p':pi,'turma':lab(k),'nome':n,'comp':cn,
                                           'campo':CAMPOS[(0,2,4,1,3,5,6).index(off)],'valor':round(v,2)})
                            v=min(max(v,0),10)
                        vals.append(round(v,2))
                    else: vals.append(None)
                if all(v is None for v in vals): continue
                d[str(cidx(cn,p))]=vals
            if d: a['p'][str(pi)]=d
# unifica grafias quase idênticas de disciplinas técnicas (ex.: "Orçamento e licitações" x "Orçamento e Licitação")
import unicodedata, difflib
def _k(s_): return unicodedata.normalize('NFKD',s_.lower()).encode('ascii','ignore').decode()
FT=[i for i in range(len(COMPS)) if ANOMES[CAREA[i]]=='Formação Técnica e Profissional']
merge={}; fusoes=[]
for x in range(len(FT)):
    for y in range(x+1,len(FT)):
        i,j=FT[x],FT[y]
        if i in merge or j in merge: continue
        if difflib.SequenceMatcher(None,_k(COMPS[i]),_k(COMPS[j])).ratio()>=0.88:
            merge[j]=i; fusoes.append((COMPS[j],COMPS[i]))
if merge:
    for a in recs.values():
        for pk,d in a['p'].items():
            nd={}
            for k,v in d.items():
                k2=str(merge.get(int(k),int(k)))
                if k2 in nd and all(z is None for z in v): continue
                nd[k2]=v
            a['p'][pk]=nd
alunos=[v for v in recs.values() if v['p']]
alunos.sort(key=lambda a:(a['s'],a['t'],a['n']))
turmas=sorted({a['t'] for a in alunos})
out={'comp':COMPS,'compArea':CAREA,'areasNomes':ANOMES,
     'areas':[n for n in ANOMES if ANOMES.index(n) in set(CAREA)],
     'turmas':turmas,'series':sorted({a['s'] for a in alunos}),'per':PER,
     'perComNotas':sorted({k for a in alunos for k in a['p']}),'alunos':alunos,'qual':qual,
     'semDados':[{'turma':t,'p':p} for t in turmas for p in ['1','2'] if p not in usados.get(t,set())],
     'incons':incons,'dups':dups,'roster':{lab(k):len(v) for k,v in roster.items()}}
out['fusoes']=[{'de':a,'para':b} for a,b in fusoes]
json.dump(out,open('notas.json','w',encoding='utf-8'),ensure_ascii=False,separators=(',',':'))
import os
print('alunos',len(alunos),'| json KB %.0f'%(os.path.getsize('notas.json')/1024))
print('\ncomponentes por área:')
for ai,an in enumerate(ANOMES):
    cs=[COMPS[i] for i in range(len(COMPS)) if CAREA[i]==ai]
    if cs: print('  %-32s %s'%(an,cs))
print('\nfusões de grafia:',fusoes)
print('\nnotas fora da escala (limitadas a 10):')
for x in incons: print('  P%d %-9s %-32s %-16s %-6s %s'%(x['p'],x['turma'],x['nome'][:32],x['comp'],x['campo'],x['valor']))
print('\nlinhas duplicadas (mantida a mais completa):')
for x in dups: print('  P%d %-9s %-34s linhas %s nº %s'%(x['p'],x['turma'],x['nome'][:34],x['linhas'],x['numeros']))
